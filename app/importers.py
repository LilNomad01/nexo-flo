import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Iterable, List, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class ImportedContactRow:
    name: str
    phone: str
    email: Optional[str]
    company: Optional[str]
    opt_in: bool


ALIASES = {
    "name": {"nome", "cliente", "contato", "name"},
    "phone": {"telefone", "celular", "whatsapp", "phone", "numero", "número", "fone"},
    "email": {"email", "e-mail", "mail"},
    "company": {"empresa", "companhia", "negocio", "negócio", "company"},
    "opt_in": {"opt_in", "opt-in", "optin", "consentimento", "consent", "autorizado", "autorizacao", "autorização"},
}
TRUTHY = {"1", "true", "sim", "yes", "s", "y", "opted_in", "autorizado", "autorizada", "consentido"}


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _field(header: str) -> Optional[str]:
    normalized = header.strip().lower().replace(" ", "_")
    for field, aliases in ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _rows_to_contacts(rows: Iterable[Iterable[object]], force_opt_in: bool, max_rows: int) -> List[ImportedContactRow]:
    iterator = iter(rows)
    try:
        raw_headers = next(iterator)
    except StopIteration as exc:
        raise ValueError("A planilha está vazia.") from exc
    headers = [_field(_clean(value)) for value in raw_headers]
    if "phone" not in headers:
        raise ValueError("A planilha precisa ter uma coluna Telefone, Celular, WhatsApp ou Phone.")
    contacts = []
    for index, row in enumerate(iterator, start=2):
        if index > max_rows + 1:
            raise ValueError(f"O limite por importação é de {max_rows} contatos.")
        values = list(row)
        mapped = {field: _clean(values[position]) if position < len(values) else "" for position, field in enumerate(headers) if field}
        phone = mapped.get("phone", "")
        if not phone:
            continue
        contacts.append(
            ImportedContactRow(
                name=mapped.get("name") or f"Contato {len(contacts) + 1}",
                phone=phone,
                email=mapped.get("email") or None,
                company=mapped.get("company") or None,
                opt_in=force_opt_in or mapped.get("opt_in", "").lower() in TRUTHY,
            )
        )
    if not contacts:
        raise ValueError("Nenhum telefone foi encontrado na planilha.")
    return contacts


def parse_contacts_file(filename: str, payload: bytes, force_opt_in: bool, max_rows: int) -> List[ImportedContactRow]:
    if not payload:
        raise ValueError("Selecione uma planilha para importar.")
    if len(payload) > 8 * 1024 * 1024:
        raise ValueError("O arquivo é muito grande. Use uma planilha de até 8 MB.")
    name = filename.lower()
    if name.endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
            return _rows_to_contacts(workbook.active.iter_rows(values_only=True), force_opt_in, max_rows)
        except Exception as exc:
            raise ValueError("Não consegui ler o arquivo XLSX. Verifique se ele não está corrompido.") from exc
    if name.endswith(".csv"):
        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = payload.decode("latin-1")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t,")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        return _rows_to_contacts(csv.reader(StringIO(text), dialect), force_opt_in, max_rows)
    raise ValueError("Formato não suportado. Envie um arquivo .xlsx ou .csv.")
